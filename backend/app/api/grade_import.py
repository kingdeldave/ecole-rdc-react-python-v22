from io import BytesIO
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.db.session import get_db
from app.deps import ensure_same_school, get_current_user, request_meta
from app.models import AuditAction, ClassSubject, Grade, Period, PeriodCode, RoleCode, Student, UserProfile
from app.schemas import GradeExcelImportOut
from app.services.audit import log_action
from app.services.grades import assert_can_enter_grade, upsert_grade

router = APIRouter(prefix="/grades", tags=["grades-excel"])
DIRECTION_ROLES = {RoleCode.PREFET, RoleCode.DIRECTEUR, RoleCode.SUPER_ADMIN}


def _norm(value) -> str:
    return str(value or "").strip().lower().replace("é", "e").replace("è", "e").replace("ê", "e")


def _parse_note(value):
    if value is None or str(value).strip() == "":
        return None
    raw = str(value).strip().replace(",", ".")
    if "/" in raw:
        raise ValueError("format interdit avec slash")
    return float(raw)


@router.post("/import-excel", response_model=GradeExcelImportOut)
def import_grades_excel(
    request: Request,
    class_subject_id: str,
    period_code: PeriodCode,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    class_subject = db.get(ClassSubject, class_subject_id)
    if not class_subject:
        raise HTTPException(status_code=404, detail="Cours de classe introuvable.")
    ensure_same_school(user, class_subject.school_id)
    assert_can_enter_grade(user, class_subject)
    period_state = db.query(Period).filter(Period.school_year_id == class_subject.classroom.school_year_id, Period.code == period_code).first()
    if period_state and period_state.is_closed and user.role not in DIRECTION_ROLES:
        raise HTTPException(status_code=403, detail="Cette période est clôturée. Import Excel refusé.")
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Fichier Excel .xlsx attendu.")
    wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Fichier vide.")
    header = [_norm(x) for x in rows[0]]
    matricule_idx = next((i for i, h in enumerate(header) if h in {"matricule", "mat", "id", "numero"}), 0)
    note_idx = next((i for i, h in enumerate(header) if h in {"note", "points", "point", period_code.value.lower()}), 1)
    data_rows = rows[1:] if any(header) else rows
    imported = 0
    skipped = 0
    errors: list[str] = []
    max_value = {
        PeriodCode.P1: class_subject.max_p1,
        PeriodCode.P2: class_subject.max_p2,
        PeriodCode.EX1: class_subject.max_ex1,
        PeriodCode.P3: class_subject.max_p3,
        PeriodCode.P4: class_subject.max_p4,
        PeriodCode.EX2: class_subject.max_ex2,
        PeriodCode.RATTRAPAGE: class_subject.max_rattrapage,
        PeriodCode.TENASOP: class_subject.max_tenasop,
        PeriodCode.BAC: class_subject.max_bac,
    }[period_code]
    for line_no, row in enumerate(data_rows, start=2):
        try:
            matricule = str(row[matricule_idx] or "").strip()
            note = _parse_note(row[note_idx])
            if not matricule or note is None:
                skipped += 1
                continue
            if note < 0 or note > max_value:
                raise ValueError(f"note {note} hors maximum {max_value}")
            student = db.query(Student).filter(Student.school_id == class_subject.school_id, Student.class_id == class_subject.class_id, Student.matricule == matricule).first()
            if not student:
                skipped += 1
                errors.append(f"Ligne {line_no}: matricule introuvable ({matricule})")
                continue
            existing = db.query(Grade).filter(Grade.student_id == student.id, Grade.class_subject_id == class_subject.id, Grade.period_code == period_code).first()
            if user.role == RoleCode.ENSEIGNANT and existing and existing.locked:
                skipped += 1
                errors.append(f"Ligne {line_no}: note verrouillée ({matricule})")
                continue
            upsert_grade(db, user=user, student_id=student.id, class_subject=class_subject, period_code=period_code, value=note, reason="Import Excel professeur")
            imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Ligne {line_no}: {exc}")
    log_action(db, user=user, action=AuditAction.IMPORT_GRADES_EXCEL, entity_type="class_subject", entity_id=str(class_subject.id), new_value={"period": period_code.value, "imported": imported, "skipped": skipped}, **request_meta(request))
    db.commit()
    return GradeExcelImportOut(message="Import Excel des points terminé.", imported=imported, skipped=skipped, errors=errors[:50])

@router.post("/import-excel-full", response_model=GradeExcelImportOut)
def import_grades_excel_full(
    request: Request,
    class_subject_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserProfile = Depends(get_current_user),
):
    """Importe un fichier Excel contenant plusieurs colonnes de points pour un cours.

    Colonnes acceptées : matricule, p1, p2, ex1, p3, p4, ex2, rattrapage, tenasop, bac.
    Chaque colonne est contrôlée avec le maximum réel du cours, comme dans le fichier
    bulletin Excel : un 11/20 reste 11 et ne devient pas 10.
    """
    class_subject = db.get(ClassSubject, class_subject_id)
    if not class_subject:
        raise HTTPException(status_code=404, detail="Cours de classe introuvable.")
    ensure_same_school(user, class_subject.school_id)
    assert_can_enter_grade(user, class_subject)
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Fichier Excel .xlsx attendu.")
    wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Fichier vide.")
    header = [_norm(x).replace(" ", "_").replace("1ere", "p1").replace("2eme", "p2") for x in rows[0]]
    matricule_idx = next((i for i, h in enumerate(header) if h in {"matricule", "mat", "id", "numero"}), 0)
    period_columns = {
        PeriodCode.P1: {"p1", "periode1", "1_periode", "premiere_periode"},
        PeriodCode.P2: {"p2", "periode2", "2_periode", "deuxieme_periode"},
        PeriodCode.EX1: {"ex1", "exam_s1", "examen_s1", "examen1"},
        PeriodCode.P3: {"p3", "periode3", "3_periode"},
        PeriodCode.P4: {"p4", "periode4", "4_periode"},
        PeriodCode.EX2: {"ex2", "exam_s2", "examen_s2", "examen2"},
        PeriodCode.RATTRAPAGE: {"rattrapage", "session_rattrapage"},
        PeriodCode.TENASOP: {"tenasop"},
        PeriodCode.BAC: {"bac", "exetat", "bac_exetat"},
    }
    idx_by_period: dict[PeriodCode, int] = {}
    for code, aliases in period_columns.items():
        idx = next((i for i, h in enumerate(header) if h in aliases), None)
        if idx is not None:
            idx_by_period[code] = idx
    if not idx_by_period:
        raise HTTPException(status_code=400, detail="Aucune colonne de points trouvée. Utilise p1, p2, ex1, p3, p4, ex2.")
    imported = 0
    skipped = 0
    errors: list[str] = []
    seen: set[str] = set()
    max_by_period = {
        PeriodCode.P1: class_subject.max_p1,
        PeriodCode.P2: class_subject.max_p2,
        PeriodCode.EX1: class_subject.max_ex1,
        PeriodCode.P3: class_subject.max_p3,
        PeriodCode.P4: class_subject.max_p4,
        PeriodCode.EX2: class_subject.max_ex2,
        PeriodCode.RATTRAPAGE: class_subject.max_rattrapage,
        PeriodCode.TENASOP: class_subject.max_tenasop,
        PeriodCode.BAC: class_subject.max_bac,
    }
    for line_no, row in enumerate(rows[1:], start=2):
        matricule = str(row[matricule_idx] or "").strip()
        if not matricule:
            skipped += 1
            continue
        if matricule in seen:
            skipped += 1
            errors.append(f"Ligne {line_no}: matricule dupliqué ({matricule})")
            continue
        seen.add(matricule)
        student = db.query(Student).filter(Student.school_id == class_subject.school_id, Student.class_id == class_subject.class_id, Student.matricule == matricule).first()
        if not student:
            skipped += 1
            errors.append(f"Ligne {line_no}: matricule introuvable ({matricule})")
            continue
        for code, idx in idx_by_period.items():
            try:
                note = _parse_note(row[idx])
                if note is None:
                    continue
                if note < 0 or note > max_by_period[code]:
                    raise ValueError(f"note {note} hors maximum {max_by_period[code]} pour {code.value}")
                existing = db.query(Grade).filter(Grade.student_id == student.id, Grade.class_subject_id == class_subject.id, Grade.period_code == code).first()
                if user.role == RoleCode.ENSEIGNANT and existing and existing.locked:
                    errors.append(f"Ligne {line_no}: note verrouillée ({matricule}, {code.value})")
                    skipped += 1
                    continue
                upsert_grade(db, user=user, student_id=student.id, class_subject=class_subject, period_code=code, value=note, reason="Import Excel complet professeur")
                imported += 1
            except Exception as exc:
                skipped += 1
                errors.append(f"Ligne {line_no} {code.value}: {exc}")
    log_action(db, user=user, action=AuditAction.IMPORT_FULL_GRADES_EXCEL, entity_type="class_subject", entity_id=str(class_subject.id), new_value={"imported": imported, "skipped": skipped, "columns": [c.value for c in idx_by_period]}, **request_meta(request))
    db.commit()
    return GradeExcelImportOut(message="Import Excel complet terminé.", imported=imported, skipped=skipped, errors=errors[:80])
