from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
from uuid import uuid4
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from app.core.config import settings
from app.db.session import get_db
from app.deps import ensure_same_school, get_current_user, request_meta
from app.models import AuditAction, ClassRoom, ClassSubject, Grade, PeriodCode, RoleCode, Student, StudentFeeStatus, UserProfile
from app.services.audit import log_action
from app.services.grades import assert_can_enter_grade, upsert_grade

router = APIRouter(prefix="/excel", tags=["excel"])




def _norm_header(value: object) -> str:
    """Normalise un titre de colonne venant d'Excel."""
    raw = str(value or "").strip().lower()
    replacements = {"é": "e", "è": "e", "ê": "e", "à": "a", "ù": "u", "ï": "i", "î": "i", "ô": "o", "ç": "c", "-": "_", " ": "_"}
    for src, dst in replacements.items():
        raw = raw.replace(src, dst)
    while "__" in raw:
        raw = raw.replace("__", "_")
    return raw.strip("_")


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_sex(value: object) -> str:
    raw = str(value or "M").strip().upper()
    if raw in {"F", "FEMININ", "FÉMININ", "FILLE", "FEMALE"}:
        return "F"
    return "M"


def _generated_path(name: str) -> Path:
    out_dir = Path(settings.GENERATED_DIR) / "excel"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / name


@router.get("/grades/export")
def export_grades(class_id: str, db: Session = Depends(get_db), user: UserProfile = Depends(get_current_user)):
    """Exporte une matrice de points au format Excel.

    Le fichier contient une feuille d'information pour tracer l'utilisateur qui a exporté.
    En production, il faut placer ce fichier dans un bucket privé et le faire expirer.
    """
    classroom = db.get(ClassRoom, class_id)
    if not classroom:
        raise HTTPException(status_code=404, detail="Classe introuvable.")
    ensure_same_school(user, classroom.school_id)

    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.last_name.asc()).all()
    class_subjects = db.query(ClassSubject).filter(ClassSubject.class_id == class_id).order_by(ClassSubject.display_order.asc()).all()
    grades = db.query(Grade).join(Student, Student.id == Grade.student_id).filter(Student.class_id == class_id).all()
    grade_map = {(g.student_id, g.class_subject_id, g.period_code.value): g.value for g in grades}

    wb = Workbook()
    ws = wb.active
    ws.title = "Points"
    info = wb.create_sheet("Informations")

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")

    base_headers = ["Matricule", "Nom complet"]
    period_codes = ["P1", "P2", "EX1", "P3", "P4", "EX2"]
    headers = base_headers + [f"{cs.subject.name} - {p}" for cs in class_subjects for p in period_codes]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for student in students:
        row = [student.matricule, student.full_name]
        for cs in class_subjects:
            for p in period_codes:
                row.append(grade_map.get((student.id, cs.id, p), ""))
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(col[0].value or "")) + 2, 12), 28)

    info.append(["Champ", "Valeur"])
    info.append(["Généré par", user.full_name])
    info.append(["Rôle", user.role.value])
    info.append(["Classe", classroom.name])
    info.append(["École", classroom.school.name])
    info.append(["Note", "Document scolaire confidentiel. Export réservé aux personnes autorisées."])

    path = _generated_path(f"points-{classroom.id}-{uuid4().hex}.xlsx")
    wb.save(path)
    log_action(db, user=user, action=AuditAction.EXPORT_EXCEL, entity_type="excel_export", entity_id=str(classroom.id), new_value={"class_id": str(classroom.id)})
    db.commit()
    return FileResponse(str(path), filename=path.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/students/import")
async def import_students(class_id: str, file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db), user: UserProfile = Depends(get_current_user)):
    """Importe des élèves depuis un fichier Excel flexible.

    Colonnes acceptées : matricule, nom, post_nom, prenom, sexe, date_naissance,
    lieu_naissance, adresse. Les accents et variantes simples sont tolérés
    (exemple : Post-nom, post nom, Date naissance).
    """
    classroom = db.get(ClassRoom, class_id)
    if not classroom:
        raise HTTPException(status_code=404, detail="Classe introuvable.")
    ensure_same_school(user, classroom.school_id)

    tmp = _generated_path(f"import-students-{uuid4().hex}.xlsx")
    tmp.write_bytes(await file.read())
    wb = load_workbook(tmp)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=422, detail="Le fichier Excel est vide.")

    headers = [_norm_header(v) for v in rows[0]]
    aliases = {
        "matricule": ["matricule", "mat", "id", "code", "numero_matricule"],
        "nom": ["nom", "last_name", "nom_de_famille"],
        "post_nom": ["post_nom", "postnom", "post", "middle_name"],
        "prenom": ["prenom", "first_name"],
        "sexe": ["sexe", "sex", "genre"],
        "date_naissance": ["date_naissance", "date_de_naissance", "naissance"],
        "lieu_naissance": ["lieu_naissance", "lieu_de_naissance"],
        "adresse": ["adresse", "address"],
    }

    def idx(name: str) -> int | None:
        for alias in aliases[name]:
            if alias in headers:
                return headers.index(alias)
        return None

    i_matricule = idx("matricule")
    i_nom = idx("nom")
    if i_matricule is None or i_nom is None:
        raise HTTPException(status_code=422, detail="Le fichier doit contenir au minimum les colonnes matricule et nom.")

    created = 0
    skipped = 0
    errors: list[str] = []
    seen: set[str] = set()

    for line_no, row in enumerate(rows[1:], start=2):
        values = list(row)
        matricule = _cell_text(values[i_matricule] if i_matricule < len(values) else None)
        nom = _cell_text(values[i_nom] if i_nom < len(values) else None)
        if not matricule or not nom:
            skipped += 1
            continue
        if matricule in seen:
            skipped += 1
            errors.append(f"Ligne {line_no}: matricule dupliqué dans le fichier ({matricule}).")
            continue
        seen.add(matricule)

        exists = db.query(Student).filter(Student.school_id == classroom.school_id, Student.matricule == matricule).first()
        if exists:
            skipped += 1
            errors.append(f"Ligne {line_no}: matricule déjà existant ({matricule}).")
            continue

        def get(name: str) -> object:
            pos = idx(name)
            if pos is None or pos >= len(values):
                return None
            return values[pos]

        student = Student(
            school_id=classroom.school_id,
            class_id=classroom.id,
            matricule=matricule,
            last_name=nom,
            middle_name=_cell_text(get("post_nom")),
            first_name=_cell_text(get("prenom")),
            sex=_normalize_sex(get("sexe")),
            birth_date=_cell_date(get("date_naissance")),
            birth_place=_cell_text(get("lieu_naissance")),
            address=_cell_text(get("adresse")),
        )
        db.add(student)
        db.flush()
        db.add(StudentFeeStatus(
            school_id=student.school_id,
            student_id=student.id,
            school_year_id=classroom.school_year_id,
            total_due=0,
            total_paid=0,
        ))
        created += 1

    log_action(db, user=user, action=AuditAction.IMPORT_EXCEL, entity_type="excel_import_students", entity_id=str(classroom.id), new_value={"created": created, "skipped": skipped, "errors": errors[:20]}, **request_meta(request))
    db.commit()
    return {"message": "Import terminé.", "created": created, "skipped": skipped, "errors": errors}


@router.post("/grades/import")
async def import_grades(class_subject_id: str, period_code: PeriodCode, file: UploadFile = File(...), request: Request = None, db: Session = Depends(get_db), user: UserProfile = Depends(get_current_user)):
    """Importe des points depuis Excel.

    Colonnes attendues : matricule, note.
    """
    class_subject = db.get(ClassSubject, class_subject_id)
    if not class_subject:
        raise HTTPException(status_code=404, detail="Cours de classe introuvable.")
    ensure_same_school(user, class_subject.school_id)
    assert_can_enter_grade(user, class_subject)

    tmp = _generated_path(f"import-grades-{uuid4().hex}.xlsx")
    tmp.write_bytes(await file.read())
    wb = load_workbook(tmp)
    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        matricule, value = (list(row) + [None] * 2)[:2]
        if not matricule or value is None:
            continue
        if isinstance(value, str) and "/" in value:
            raise HTTPException(status_code=422, detail="Une note contient le format interdit 93/80.")
        student = db.query(Student).filter(Student.school_id == class_subject.school_id, Student.matricule == str(matricule), Student.class_id == class_subject.class_id).first()
        if not student:
            continue
        upsert_grade(db, user=user, student_id=student.id, class_subject=class_subject, period_code=period_code, value=float(value), reason="Import Excel")
        imported += 1
    log_action(db, user=user, action=AuditAction.IMPORT_EXCEL, entity_type="excel_import", entity_id=str(class_subject.id), new_value={"imported": imported}, **request_meta(request))
    db.commit()
    return {"message": "Import des points terminé.", "imported": imported}
